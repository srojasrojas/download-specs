#!/usr/bin/env python3
"""
Analyze duplicate URLs in the most recent exported_specs*.xlsx.

A URL is considered "duplicate" when it appears in more than one row.
The script reports which columns differ between those rows and exports
a CSV with the full detail of every duplicate group.

Usage:
    python analyze_dupes.py
    python analyze_dupes.py --out dupes_report.csv
"""

import argparse
import csv
import glob
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

SPEC_URL_COLUMN = "specification"
HEADER_ROW = 2
DATA_START_ROW = 3


def find_latest_excel() -> Path:
    matches = sorted(glob.glob("exported_specs*.xlsx"))
    if not matches:
        sys.exit("ERROR: No exported_specs*.xlsx file found in the current directory.")
    return Path(matches[-1])


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Report duplicate spec URLs in exported_specs*.xlsx."
    )
    parser.add_argument(
        "--out",
        metavar="FILE",
        default="dupes_report.csv",
        help="Output CSV path (default: dupes_report.csv)",
    )
    args = parser.parse_args()

    xlsx = find_latest_excel()
    print(f"Reading {xlsx} ...")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active

    headers: list[str] = []
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW):
        headers = [str(c.value).strip() if c.value else "" for c in row]

    if SPEC_URL_COLUMN not in headers:
        wb.close()
        sys.exit(f"ERROR: Column '{SPEC_URL_COLUMN}' not found. Available: {headers}")

    url_idx = headers.index(SPEC_URL_COLUMN)

    # Group all rows by URL
    url_groups: dict[str, list[tuple]] = defaultdict(list)
    total_rows = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        url = row[url_idx] if len(row) > url_idx else None
        if not url or not str(url).startswith("http"):
            continue
        total_rows += 1
        url_groups[str(url).strip()].append(row)

    wb.close()

    dupe_groups = {url: rows for url, rows in url_groups.items() if len(rows) > 1}
    extra_rows = sum(len(v) - 1 for v in dupe_groups.values())

    # ── Console summary ──────────────────────────────────────────────────────
    print()
    print("=== Deduplication summary ===")
    print(f"  Total data rows      : {total_rows:,}")
    print(f"  Unique URLs          : {len(url_groups):,}")
    print(f"  URLs with duplicates : {len(dupe_groups):,}")
    print(f"  Extra rows (dropped) : {extra_rows:,}")
    print()

    # Which columns differ inside duplicate groups?
    col_diffs: dict[str, int] = defaultdict(int)
    for rows in dupe_groups.values():
        for col_i, col_name in enumerate(headers):
            if not col_name:
                continue
            vals = {r[col_i] if len(r) > col_i else None for r in rows}
            if len(vals) > 1:
                col_diffs[col_name] += 1

    print("  Columns that differ between rows sharing the same URL:")
    for col, count in sorted(col_diffs.items(), key=lambda x: -x[1]):
        pct = count / len(dupe_groups) * 100
        print(f"    {col:<22} {count:>6,} / {len(dupe_groups):,} groups  ({pct:.1f}%)")

    # ── CSV export ───────────────────────────────────────────────────────────
    out_path = Path(args.out)
    csv_headers = ["url", "occurrences", "differing_columns"] + headers

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(csv_headers)
        for url, rows in dupe_groups.items():
            # Find which columns differ in this specific group
            differing = []
            for col_i, col_name in enumerate(headers):
                if not col_name:
                    continue
                vals = {r[col_i] if len(r) > col_i else None for r in rows}
                if len(vals) > 1:
                    differing.append(col_name)

            for row in rows:
                writer.writerow(
                    [url, len(rows), "|".join(differing)] + list(row)
                )

    print()
    print(f"Full detail exported to: {out_path}")
    print(f"  Rows in CSV: {sum(len(v) for v in dupe_groups.values()):,}")


if __name__ == "__main__":
    main()
