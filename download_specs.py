#!/usr/bin/env python3
"""
Download product specification files listed in the most recent exported_specs*.xlsx.

Usage:
    python download_specs.py
    python download_specs.py --subfolder countryName
    python download_specs.py --subfolder providerName --delay 3 --output specs
    python download_specs.py --test 10
"""

import argparse
import glob
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import requests

# ── Constants ────────────────────────────────────────────────────────────────
SPEC_URL_COLUMN = "specification"
HEADER_ROW = 2       # Row 1 is blank; headers are on row 2
DATA_START_ROW = 3
DEFAULT_OUTPUT = "specs"
DEFAULT_DELAY = 2.0  # seconds between downloads
REQUEST_TIMEOUT = 30  # seconds per request


# ── Logger ───────────────────────────────────────────────────────────────────

class Logger:
    """Writes simultaneously to stdout and a log file."""

    def __init__(self, log_path: Path) -> None:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        self._file = open(log_path, "w", encoding="utf-8")
        self.path = log_path

    def log(self, msg: str = "", *, file_only: bool = False) -> None:
        self._file.write(msg + "\n")
        self._file.flush()
        if not file_only:
            print(msg)

    def close(self) -> None:
        self._file.close()


# ── Helpers ──────────────────────────────────────────────────────────────────

def find_latest_excel() -> Path:
    matches = sorted(glob.glob("exported_specs*.xlsx"))
    if not matches:
        sys.exit("ERROR: No exported_specs*.xlsx file found in the current directory.")
    return Path(matches[-1])


def fmt_duration(seconds: float) -> str:
    if seconds < 60:
        return f"{seconds:.1f}s"
    return str(timedelta(seconds=int(seconds)))


def read_rows(xlsx_path: Path, subfolder_col: str | None) -> tuple[list[dict], dict]:
    """
    Return (rows, dedup_stats).

    rows: list of dicts with keys: url, subfolder, filename.
    dedup_stats: summary of deduplication findings.
    """
    print(f"Reading {xlsx_path} ...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    headers = []
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW):
        headers = [str(c.value).strip() if c.value else "" for c in row]

    if SPEC_URL_COLUMN not in headers:
        wb.close()
        sys.exit(f"ERROR: Column '{SPEC_URL_COLUMN}' not found. Available: {headers}")
    if subfolder_col and subfolder_col not in headers:
        wb.close()
        sys.exit(f"ERROR: Column '{subfolder_col}' not found. Available: {headers}")

    url_idx = headers.index(SPEC_URL_COLUMN)
    sub_idx = headers.index(subfolder_col) if subfolder_col else None

    # url -> list of raw rows (for dedup analysis)
    url_groups: dict[str, list[tuple]] = defaultdict(list)
    total_data_rows = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        url = row[url_idx] if len(row) > url_idx else None
        if not url or not str(url).startswith("http"):
            continue
        total_data_rows += 1
        url_groups[str(url).strip()].append(row)

    wb.close()

    # Build dedup stats
    dupe_groups = {url: rows for url, rows in url_groups.items() if len(rows) > 1}
    extra_rows = sum(len(v) - 1 for v in dupe_groups.values())

    # Find which columns differ within duplicate groups
    col_diffs: dict[str, int] = defaultdict(int)
    for rows in dupe_groups.values():
        for col_i, col_name in enumerate(headers):
            if not col_name:
                continue
            vals = {r[col_i] if len(r) > col_i else None for r in rows}
            if len(vals) > 1:
                col_diffs[col_name] += 1

    dedup_stats = {
        "total_data_rows": total_data_rows,
        "unique_urls": len(url_groups),
        "dupe_url_count": len(dupe_groups),
        "extra_rows_dropped": extra_rows,
        "cols_that_differ": dict(
            sorted(col_diffs.items(), key=lambda x: -x[1])
        ),
    }

    # Build final row list (first occurrence per URL)
    rows_out: list[dict] = []
    for url, group_rows in url_groups.items():
        first = group_rows[0]
        subfolder = ""
        if sub_idx is not None:
            val = first[sub_idx] if len(first) > sub_idx else None
            if val:
                subfolder = "".join(
                    c if c.isalnum() or c in " _-" else "_"
                    for c in str(val).strip()
                )
        filename = Path(urlparse(url).path).name or f"spec_{len(rows_out)}.xlsx"
        rows_out.append({"url": url, "subfolder": subfolder, "filename": filename})

    return rows_out, dedup_stats


def download(url: str, dest: Path, session: requests.Session) -> tuple[bool, float]:
    """Download url to dest. Returns (success, elapsed_seconds)."""
    t0 = time.perf_counter()
    try:
        resp = session.get(url, timeout=REQUEST_TIMEOUT, stream=True)
        resp.raise_for_status()
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                f.write(chunk)
        return True, time.perf_counter() - t0
    except requests.RequestException as exc:
        elapsed = time.perf_counter() - t0
        print(f"  FAIL  {exc}", file=sys.stderr)
        return False, elapsed


# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Bulk-download product spec files from an exported_specs*.xlsx."
    )
    parser.add_argument(
        "--subfolder",
        metavar="COLUMN",
        default=None,
        help=(
            "Column whose value is used as a sub-directory inside --output. "
            "E.g. --subfolder countryName  ->  specs/Chile/Spec_xxx.xlsx"
        ),
    )
    parser.add_argument(
        "--output",
        metavar="DIR",
        default=DEFAULT_OUTPUT,
        help=f"Root output directory (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--delay",
        metavar="SECONDS",
        type=float,
        default=DEFAULT_DELAY,
        help=f"Sleep between downloads in seconds (default: {DEFAULT_DELAY})",
    )
    parser.add_argument(
        "--no-skip",
        action="store_true",
        help="Re-download files that already exist locally (default: skip them)",
    )
    parser.add_argument(
        "--test",
        metavar="N",
        type=int,
        default=None,
        help="Download only the first N specs (useful for testing)",
    )
    args = parser.parse_args()

    run_ts = datetime.now()
    xlsx = find_latest_excel()
    rows, dedup = read_rows(xlsx, args.subfolder)

    if not rows:
        sys.exit("No downloadable rows found.")

    if args.test is not None:
        rows = rows[: args.test]

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    log_path = output_dir / f"download_{run_ts.strftime('%Y%m%d_%H%M%S')}.log"
    logger = Logger(log_path)

    total = len(rows)
    skipped = downloaded = failed = 0
    download_times: list[float] = []

    # ── Header ──────────────────────────────────────────────────────────────
    logger.log(f"=== download_specs  {run_ts.strftime('%Y-%m-%d %H:%M:%S')} ===")
    logger.log(f"Source : {xlsx}")
    logger.log(f"Output : {output_dir}/")
    if args.subfolder:
        logger.log(f"Subfolder column : '{args.subfolder}'")
    logger.log(f"Delay  : {args.delay}s")
    if args.test is not None:
        logger.log(f"Mode   : TEST (first {args.test} specs)")
    logger.log()

    # ── Dedup summary ────────────────────────────────────────────────────────
    logger.log("--- Deduplication summary ---")
    logger.log(f"  Total rows in Excel  : {dedup['total_data_rows']:,}")
    logger.log(f"  Unique URLs          : {dedup['unique_urls']:,}")
    logger.log(f"  URLs with duplicates : {dedup['dupe_url_count']:,}")
    logger.log(f"  Rows dropped (dupes) : {dedup['extra_rows_dropped']:,}")
    logger.log("  Columns that differ between duplicated rows:")
    for col, count in dedup["cols_that_differ"].items():
        logger.log(f"    {col:<22} {count:,} groups")
    logger.log()

    # ── Download loop ────────────────────────────────────────────────────────
    logger.log(f"--- Downloads ({total} specs) ---")

    session = requests.Session()
    session.headers.update({"User-Agent": "spec-downloader/1.0"})
    pad = len(str(total))

    run_start = time.perf_counter()

    for i, row in enumerate(rows, start=1):
        dest = (
            output_dir / row["subfolder"] / row["filename"]
            if row["subfolder"]
            else output_dir / row["filename"]
        )
        prefix = f"[{i:>{pad}}/{total}]"

        if not args.no_skip and dest.exists():
            msg = f"{prefix} SKIP  {dest}"
            logger.log(msg)
            skipped += 1
            continue

        logger.log(f"{prefix} GET   {row['url']}")
        logger.log(f"{'':>{pad+8}}{dest}", file_only=True)

        ok, elapsed = download(row["url"], dest, session)

        if ok:
            downloaded += 1
            download_times.append(elapsed)
            status = f"OK  ({fmt_duration(elapsed)})"
        else:
            failed += 1
            status = f"FAIL ({fmt_duration(elapsed)})"

        # Terminal: show status + elapsed on same line as the GET
        print(f"{'':>{pad+8}}{dest}  [{status}]")
        logger.log(f"{'':>{pad+8}}{status}", file_only=True)

        # ETA line (only in --test mode, after at least one download)
        if args.test is not None and download_times:
            avg = sum(download_times) / len(download_times)
            full_eta = fmt_duration(avg * dedup["unique_urls"])
            logger.log(
                f"{'':>{pad+8}}avg {fmt_duration(avg)}/file  "
                f"-> projected full run: {full_eta}"
            )

        if i < total:
            time.sleep(args.delay)

    # ── Summary ──────────────────────────────────────────────────────────────
    total_elapsed = time.perf_counter() - run_start
    avg_dl = sum(download_times) / len(download_times) if download_times else 0

    logger.log()
    logger.log("--- Summary ---")
    logger.log(f"  Downloaded : {downloaded}")
    logger.log(f"  Skipped    : {skipped}")
    logger.log(f"  Failed     : {failed}")
    logger.log(f"  Total time : {fmt_duration(total_elapsed)}")
    if avg_dl:
        logger.log(f"  Avg / file : {fmt_duration(avg_dl)}")
    if args.test is not None and avg_dl:
        logger.log(
            f"  Projected full run ({dedup['unique_urls']:,} specs): "
            f"{fmt_duration(avg_dl * dedup['unique_urls'])}"
        )
    logger.log(f"\nLog written to: {log_path}")
    logger.close()


if __name__ == "__main__":
    main()
